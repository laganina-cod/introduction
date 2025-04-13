import os
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList

def create_chart(ws, title):
    """Создает scatter-график между столбцами A (X) и C (Y)"""
    chart = ScatterChart()
    chart.title = title
    chart.style = 13
    chart.x_axis.title = 'Данные столбца A'
    chart.y_axis.title = 'Данные столбца C'
    

    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()
    
 
    x_values = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)  # Столбец A
    y_values = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)  # Столбец C
    
    series = Series(y_values, x_values, title="A vs C")
    series.marker.symbol = 'circle'
    series.marker.size = 8
    
    series.dLbls = DataLabelList()
    series.dLbls.showVal = True
    series.dLbls.position = 't'
    
    chart.series.append(series)
    return chart

def add_charts_to_excel():
    """Основная функция обработки файлов"""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    target_dir = os.path.join(base_dir, 'results', 'statistics')
    
    if not os.path.exists(target_dir):
        print(f"Ошибка: папка {target_dir} не найдена!")
        return
    
    print(f"Поиск файлов в: {target_dir}")
    
    for root, _, files in os.walk(target_dir):
        for file in files:
            if not file.lower().endswith('.xlsx') or file.startswith('~$'):
                continue
                
            file_path = os.path.join(root, file)
            print(f"\nОбработка: {file_path}")
            
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                
              
                if ws.max_row < 2 or not any(cell.value for cell in ws['A']) or not any(cell.value for cell in ws['C']):
                    print(f"Пропуск: в файле {file} нет данных в столбцах A и/или C")
                    continue
                
                chart_title = f"Зависимость: A vs C\n{os.path.basename(file)}"
                main_chart = create_chart(ws, chart_title)
                
                if 'Графики' not in wb.sheetnames:
                    charts_sheet = wb.create_sheet('Графики')
                else:
                    charts_sheet = wb['Графики']
                    charts_sheet.delete_rows(1, charts_sheet.max_row)
                
                charts_sheet.add_chart(main_chart, 'A1')
                wb.save(file_path)
                print(f"Успешно: график A vs C добавлен в {file}")
                
            except PermissionError:
                print(f"Ошибка: файл {file} занят другим процессом")
            except Exception as e:
                print(f"Ошибка: {type(e).__name__} - {str(e)}")

if __name__ == "__main__":
    print("=== Построение графиков A vs C ===")
    add_charts_to_excel()
    print("\nГотово! Проверьте файлы в папке results/statistics")
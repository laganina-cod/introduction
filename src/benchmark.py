import os
import sys
import subprocess
import statistics
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# ==================================================
# КОНФИГУРАЦИЯ ПУТЕЙ (автоматически определяется)
# ==================================================
SCRIPT_DIR = Path(__file__).parent.absolute()
PROJECT_ROOT = SCRIPT_DIR.parent.absolute()
RESULTS_DIR = PROJECT_ROOT / 'results'

# Пути к исполняемым файлам (должны находиться рядом со скриптом)
EXECUTABLES = {
    'sp': SCRIPT_DIR / 'sp.exe',
    'mxv_1d': SCRIPT_DIR / 'mxv_1d.exe',
    'mxv_2d': SCRIPT_DIR / 'mxv_2d.exe',
    'mxm_1d': SCRIPT_DIR / 'mxm_1d.exe',
    'mxm_2d': SCRIPT_DIR / 'mxm_2d.exe'
}

# Соответствие data_method их описанию
DATA_METHODS = {
    1: 'zeros',
    2: 'random',
    3: 'twos',
    4: 'custom_pattern'  # Новый тип данных
}

# ==================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ==================================================
def get_results_path(method, data_method, matrix_storage, n=None, stat_type=None):
    """Генерирует правильные пути к файлам результатов"""
    if n is not None:
        # Путь для сырых данных
        return RESULTS_DIR / 'raw_data' / f"{method}_{data_method}_{matrix_storage}_{n}.xlsx"
    else:
        # Путь для статистики
        return RESULTS_DIR / 'statistics' / stat_type / f"{method}_{data_method}_{matrix_storage}.xlsx"

def create_directories():
    """Создает все необходимые директории"""
    (RESULTS_DIR / 'raw_data').mkdir(parents=True, exist_ok=True)
    (RESULTS_DIR / 'statistics' / 'mean').mkdir(parents=True, exist_ok=True)
    (RESULTS_DIR / 'statistics' / 'trimmed_mean').mkdir(parents=True, exist_ok=True)
    (RESULTS_DIR / 'statistics' / 'min').mkdir(parents=True, exist_ok=True)

def save_to_excel(filepath, data, headers=None, sheet_name="Results"):
    """
    Сохраняет данные в Excel файл (добавляет к существующим)
    """
    try:
        # Пытаемся загрузить существующий файл
        try:
            wb = load_workbook(filepath)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        except (FileNotFoundError, InvalidFileException):
            # Создаем новый файл если не существует
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
        
        # Добавляем заголовки если лист пустой
        if headers and ws.max_row == 1 and ws.max_column == 1:
            ws.append(headers)
        
        # Добавляем данные
        for row in data:
            ws.append(row)
        
        wb.save(filepath)
    except Exception as e:
        print(f"Ошибка при сохранении в {filepath}: {e}")

def run_operation(method, n, data_method, matrix_storage):
    """Запускает тестируемую операцию и возвращает время"""
    try:
        if method == 1:
            exe = EXECUTABLES['sp']
        elif method == 2:
            exe = EXECUTABLES['mxv_1d'] if matrix_storage == 0 else EXECUTABLES['mxv_2d']
        elif method == 3:
            exe = EXECUTABLES['mxm_1d'] if matrix_storage == 0 else EXECUTABLES['mxm_2d']
        
        result = subprocess.run([str(exe), str(n), str(data_method)],
                              capture_output=True, text=True, check=True)
        return float(result.stdout.strip())
    except subprocess.CalledProcessError as e:
        print(f"Ошибка выполнения {exe.name}: {e}")
        return None
    except ValueError:
        print(f"Ошибка преобразования результата для {exe.name}")
        return None

def calculate_stats(run_times):
    """Вычисляет статистические показатели"""
    if not run_times:
        return None, None, None
    
    mean = statistics.mean(run_times)
    
    # Усеченное среднее (без 20% худших результатов)
    trim_amount = int(len(run_times) * 0.2)
    trimmed_times = sorted(run_times)[:-trim_amount] if trim_amount else run_times
    trimmed_mean = statistics.mean(trimmed_times)
    
    min_time = min(run_times)
    
    return mean, trimmed_mean, min_time

# ==================================================
# ОСНОВНАЯ ФУНКЦИЯ
# ==================================================
def main():
    # Проверка аргументов командной строки
    if len(sys.argv) != 6:
        print("Использование: python benchmark.py <method> <n> <runs> <data_method> <storage>")
        print("  method: 1=sp, 2=mxv, 3=mxm")
        print("  n: размер матрицы/вектора")
        print("  runs: количество запусков")
        print(f"  data_method: 1=zeros, 2=random, 3=twos, 4=custom_pattern")
        print("  storage: 0=1D array, 1=2D array")
        sys.exit(1)
    
    try:
        method = int(sys.argv[1])
        n = int(sys.argv[2])
        num_runs = int(sys.argv[3])
        data_method = int(sys.argv[4])
        matrix_storage = int(sys.argv[5])
        
        # Валидация входных данных
        if method not in {1, 2, 3}:
            raise ValueError("Method должен быть 1, 2 или 3")
        if data_method not in {1, 2, 3, 4}:  # Добавлена проверка для data_method=4
            raise ValueError("Data_method должен быть 1, 2, 3 или 4")
        if matrix_storage not in {0, 1}:
            raise ValueError("Storage должен быть 0 или 1")
        
        # Подготовка директорий
        create_directories()
        
        # Запуск тестов
        run_times = []
        for _ in range(num_runs):
            time = run_operation(method, n, data_method, matrix_storage)
            if time is not None:
                run_times.append(time)
        
        if not run_times:
            print("Нет валидных результатов для сохранения")
            return
        
        # Сохранение сырых данных
        raw_file = get_results_path(method, data_method, matrix_storage, n)
        save_to_excel(raw_file, 
                     [[i+1, t] for i, t in enumerate(run_times)],
                     ['Run', 'Time (s)'])
        
        # Расчет и сохранение статистики
        mean, trimmed_mean, min_time = calculate_stats(run_times)
        
        stat_types = {
            'mean': mean,
            'trimmed_mean': trimmed_mean,
            'min': min_time
        }
        
        for stat_name, value in stat_types.items():
            stat_file = get_results_path(method, data_method, matrix_storage, stat_type=stat_name)
            save_to_excel(stat_file, [[n, value]], ['Size', 'Time (s)'])
        
        data_desc = DATA_METHODS.get(data_method, f'unknown({data_method})')
        print(f"Тестирование завершено. Метод: {method}, data: {data_desc}, storage: {matrix_storage}")
        print(f"Результаты сохранены в {RESULTS_DIR}")
    
    except Exception as e:
        print(f"Ошибка: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
import os

def create_chart(ws, title):
    """Создает полностью настроенный scatter-график"""
    chart = ScatterChart()
    
    # Основные настройки
    chart.title = title
    chart.style = 13  # Стиль оформления
    chart.x_axis.title = 'Размер матрицы (N)'
    chart.y_axis.title = 'Время выполнения (сек)'
    
    # Настройка осей
    chart.x_axis.majorGridlines = ChartLines()
    chart.y_axis.majorGridlines = ChartLines()
    
    # Данные для графика
    x_values = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    y_values = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    
    # Создание серии
    series = Series(y_values, x_values, title="Данные")
    series.marker.symbol = 'circle'  # Тип маркера
    series.marker.size = 8           # Размер маркеров
    
    # Настройка подписей
    series.dLbls = DataLabelList()
    series.dLbls.showVal = True      # Показать значения
    series.dLbls.position = 't'      # Позиция подписей (top)
    
    chart.series.append(series)
    return chart

def add_charts_to_excel():
    """Добавляет графики во все Excel-файлы"""
    for root, _, files in os.walk('results/statistics'):
        for file in files:
            if not file.endswith('.xlsx'):
                continue
                
            file_path = os.path.join(root, file)
            print(f"Обработка: {file_path}")
            
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                
                # Проверка данных
                if ws.max_row < 2:
                    print(f"Недостаточно данных в {file}")
                    continue
                
                # Создаем основной график
                main_chart = create_chart(ws, f"Зависимость времени от размера\n{file}")
                
                # Создаем лист для графиков (если нет)
                if 'Графики' not in wb.sheetnames:
                    charts_sheet = wb.create_sheet('Графики')
                else:
                    charts_sheet = wb['Графики']
                    charts_sheet.delete_rows(1, 100)  # Очищаем старые графики
                
                # Добавляем графики на лист
                charts_sheet.add_chart(main_chart, 'A1')
                
                # Сохраняем изменения
                wb.save(file_path)
                print(f"Графики добавлены в {file}")
                
            except Exception as e:
                print(f"Ошибка при обработке {file}: {str(e)}")

if __name__ == "__main__":
    print("Добавление графиков в файлы результатов...")
    add_charts_to_excel()
    print("Готово! Проверьте файлы в папке results/statistics")
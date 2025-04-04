import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import tempfile

def add_charts_to_excel_files():
    # Настройки
    input_dir = Path('results')  # Корневая директория для поиска
    processed_files = 0

    # Сбор всех XLSX-файлов
    for root, _, files in os.walk(input_dir):
        for file in files:
            if not file.endswith('.xlsx'):
                continue

            file_path = Path(root) / file
            print(f"Обработка файла: {file_path}")

            try:
                # Открываем книгу в режиме записи
                wb = load_workbook(file_path)
                ws_data = wb.active
                
                # Собираем данные из столбцов A (размер) и C (значения)
                sizes = []
                values = []
                
                for row in ws_data.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None and row[2] is not None:
                        try:
                            sizes.append(float(row[0]))
                            values.append(float(row[2]))
                        except (TypeError, ValueError):
                            continue
                
                if not sizes:
                    print(f"Пропуск файла {file}: нет данных в столбцах A и C")
                    continue

                # Создаем график matplotlib
                plt.figure(figsize=(10, 6))
                plt.plot(sizes, values, 'bo-', linewidth=1, markersize=4)
                plt.title(f'Зависимость значений от размерности\n{file}')
                plt.xlabel('Размерность (N)')
                plt.ylabel('Значение из столбца C')
                plt.grid(True)
                
                # Сохраняем график во временный файл
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                    temp_img_path = tmp.name
                    plt.savefig(temp_img_path, dpi=150, bbox_inches='tight')
                    plt.close()
                
                # Создаем лист для графиков (или очищаем существующий)
                if 'Графики' in wb.sheetnames:
                    ws_chart = wb['Графики']
                    wb.remove(ws_chart)
                ws_chart = wb.create_sheet('Графики')
                
                # Вставляем изображение графика
                img = Image(temp_img_path)
                ws_chart.add_image(img, 'A1')
                
                # Создаем интерактивный график OpenPyXL
                chart = ScatterChart()
                chart.title = f"Зависимость значений от размерности ({file})"
                chart.x_axis.title = "Размерность (N)"
                chart.y_axis.title = "Значение из столбца C"
                
                x_values = Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)
                y_values = Reference(ws_data, min_col=3, min_row=2, max_row=ws_data.max_row)
                
                series = Series(y_values, x_values, title="Данные")
                chart.series.append(series)
                
                # Размещаем график на листе
                ws_chart.add_chart(chart, 'A20')
                
                # Сохраняем изменения
                wb.save(file_path)
                processed_files += 1
                print(f"Графики добавлены в {file}")
                
                # Удаляем временный файл
                os.unlink(temp_img_path)
                
            except Exception as e:
                print(f"Ошибка обработки {file}: {str(e)}")

    print(f"\nОбработка завершена. Обновлено файлов: {processed_files}")

if __name__ == "__main__":
    print("Добавление графиков в Excel-файлы...")
    add_charts_to_excel_files()